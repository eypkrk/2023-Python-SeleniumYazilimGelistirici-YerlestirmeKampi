from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from datetime import time
from pathlib import Path
from time import sleep
from datetime import date
import openpyxl
import pytest
from constants import globalConstant

class TestAllSauce:
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        #url Global olarak tanımlandı
        self.driver.get(globalConstant.URL)
        #klasör Oluşturma işlemi
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)

    def teardown_method(self):
        self.driver.quit()

        
    #excelden veri okuma
    def getUsers():
        excelFile = openpyxl.load_workbook("data/invalid_login1.xlsx")
        selectedSheet = excelFile["Sheet1"]
        totalRows = selectedSheet.max_row
        data = []
        for i in range(2,totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
        return data
    
    def waitArea(self,locater):
        WebDriverWait(self.driver,3).until(ec.visibility_of_element_located((locater)))
   
    #başarılı giriş işlemi
    def test_invalidlogin(self):
        WebDriverWait(self.driver,3).until(ec.visibility_of_element_located((By.ID,"user-name")))
        userName = self.driver.find_element(By.ID,"user-name")
        userName.send_keys("standard_user")
        WebDriverWait(self.driver,3).until(ec.visibility_of_element_located((By.ID,"password")))
        password = self.driver.find_element(By.ID,"password")
        password.send_keys("secret_sauce")
        WebDriverWait(self.driver,3).until(ec.visibility_of_element_located((By.ID,"login-button")))
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/testinvalidLogin-{userName}-{password}.png")
        self.driver.execute_script("window.scrollTo(0,500)")
        siteMsg = self.driver.find_element(By.XPATH,"//*[@id='page_wrapper']/footer/div")
        assert siteMsg.text == "© 2023 Sauce Labs. All Rights Reserved. Terms of Service | Privacy Policy" 
        
    #çift kullanıcı göndere işlemi
    @pytest.mark.parametrize("username,password",[("deneme","deneme"),("test","test")])
    def test_validlogin(self,username,password):
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"user-name")))
        userName = self.driver.find_element(By.ID,"user-name")
        userName.send_keys(username)
        WebDriverWait(self.driver,5).until(ec.visibility_of_all_elements_located((By.ID,"password")))
        passWord = self.driver.find_element(By.ID,"password")
        passWord.send_keys(password)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/testvalidLogin-{username}-{password}.png")
        WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//div[@id='login_button_container']/div/form/div[3]/h3")))
        errorMessage = self.driver.find_element(By.XPATH,"//div[@id='login_button_container']/div/form/div[3]/h3")
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    #kullanıcı parola boş bırktığında
    def test_nullPasswordArea(self):
        WebDriverWait(self.driver,3).until(ec.visibility_of_element_located((By.ID,"user-name")))
        username = self.driver.find_element(By.ID,"user-name")
        username.send_keys("kullanıcıadı")
        WebDriverWait(self.driver,3).until(ec.visibility_of_element_located((By.ID,"password")))
        password = self.driver.find_element(By.ID,"password")
        password.send_keys("")
        WebDriverWait(self.driver,3).until(ec.visibility_of_element_located((By.ID,"login-button")))
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/testnullPassword-{username}-{password}.png")
        errorMsg = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]")
        assert errorMsg.text == "Epic sadface: Password is required"
        closeBtn = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3/button")
        closeBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/testnullPasswordc-{username}-{password}.png")
        assert errorMsg.is_displayed()
    #excel den veri okuma işlemi
    @pytest.mark.parametrize("username,password",getUsers())
    def test_readUsers(self,username,password):
        self.waitArea((By.ID,"user-name"))
        userName = self.driver.find_element(By.ID,"user-name")
        userName.send_keys(username)
        self.waitArea((By.ID,"password"))
        passWord = self.driver.find_element(By.ID,"password")
        passWord.send_keys(password)
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/testreadUsers-{username}-{password}.png")
        errorMsg = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]")
        assert errorMsg.text == "Epic sadface: Username and password do not match any user in this service"
    #genel site testi
    def test_surfinSite(self):
        self.waitArea((By.ID,"user-name"))
        username = self.driver.find_element(By.ID,"user-name")
        username.send_keys("standard_user")
        password = self.driver.find_element(By.ID,"password")
        password.send_keys("secret_sauce")
        loginBtn = self.driver.find_element(By.ID,"login-button")
        loginBtn.click()
        filterBtn = self.driver.find_element(By.XPATH,"//*[@id='header_container']/div[2]/div/span/select")
        filterBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/filter.png")
        self.waitArea((By.XPATH,"//div[@id='header_container']/div[2]/div/span/select"))
        sortAsc = self.driver.find_element(By.XPATH,"//*[@id='header_container']/div[2]/div/span/select/option[2]")
        sortAsc.click()
        addCartBtn = self.driver.find_element(By.ID,"add-to-cart-test.allthethings()-t-shirt-(red)")
        addCartBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/add.png")
        basketBtn = self.driver.find_element(By.ID,"shopping_cart_container")
        basketBtn.click()
        productTxt = self.driver.find_element(By.ID,"item_3_title_link")
        assert productTxt.text == "Test.allTheThings() T-Shirt (Red)"
        productTxt.click()
        oneIcon = self.driver.find_element(By.XPATH,"//*[@id='shopping_cart_container']/a/span")
        assert oneIcon.is_enabled()
        self.waitArea((By.ID,"remove-test.allthethings()-t-shirt-(red)"))
        removeBtn = self.driver.find_element(By.ID,"remove-test.allthethings()-t-shirt-(red)")
        removeBtn.click()
        backBtn = self.driver.find_element(By.ID,"back-to-products")
        backBtn.click()
        countProducts = self.driver.find_elements(By.CLASS_NAME,"inventory_item")
        assert len(countProducts)==6
        menuBtn = self.driver.find_element(By.ID,"react-burger-menu-btn")
        menuBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/menu.png")
        self.waitArea((By.ID,"logout_sidebar_link"))
        extBtn = self.driver.find_element(By.ID,"logout_sidebar_link")
        extBtn.click()


        
    